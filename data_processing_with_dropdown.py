import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Dropdown options for categories
dropdown_options = [
    'personal care-deodorant',
    'personal care-perfume',
    'personal care-personal care (misc)',
    'personal care-beard care',
    'personal care-oral care',
    'personal care-toothpaste',
    'Other category',
    'Out of scope'
]

def add_dropdown(ws, chunk):
    """
    Adds a dropdown list to the Excel worksheet and writes data from a chunk.

    Args:
        ws (openpyxl worksheet): The worksheet where data will be added.
        chunk (pandas DataFrame): The chunk of data to be added to the worksheet.
    """
    # Add headers if the worksheet is empty
    if not ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3, values_only=True):
        ws.append(['sentence', 'category', 'main_category'])

    # Write row data for the chunk
    for _, row in chunk.iterrows():
        ws.append([row['sentence'], row['category'], row['main_category']])

    # Create a data validation (dropdown) object
    data_validation = DataValidation(type="list", formula1=f'"{",".join(dropdown_options)}"', allow_blank=True)

    # Apply the data validation to Column C (main_category)
    for row in range(2, len(chunk) + 2):  # Start from row 2, headers are in row 1
        cell = ws.cell(row=row, column=3)  # Column C is index 3
        data_validation.add(cell)

    # Add the data validation to the worksheet
    ws.add_data_validation(data_validation)

def split_and_process_chunks(df, chunk_size, output_files):
    """
    Splits the dataframe into chunks and processes each chunk into separate Excel files.
    
    Args:
        df (pandas DataFrame): The input dataframe to be split.
        chunk_size (int): The size of each chunk.
        output_files (list): A list of output file names for each chunk.
    """
    num_chunks = len(df) // chunk_size + 1

    # Create the output folder if it doesn't exist
    os.makedirs('food_data', exist_ok=True)

    # Process each chunk
    for i in range(num_chunks):
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, len(df))  # Handle the last chunk
        chunk = df.iloc[start_idx:end_idx]
        intern_name = output_files[i % len(output_files)]  # Cycle through output files

        # Create a new Excel workbook for each chunk
        wb = Workbook()
        ws = wb.active
        output_file = f"food_data/{intern_name}.xlsx"

        # Add dropdown and data to the worksheet
        add_dropdown(ws, chunk)

        # Save the workbook
        wb.save(output_file)

def count_sentences_in_files(output_folder, output_file):
    """
    Counts the number of sentences in each CSV file in the given folder and writes the results to a CSV file.

    Args:
        output_folder (str): The folder where CSV files are stored.
        output_file (str): The file where results will be saved.
    """
    if not os.path.exists(output_folder):
        print(f"The folder {output_folder} does not exist.")
        return

    # Get a list of all CSV files in the output folder
    files = [f for f in os.listdir(output_folder) if f.endswith('.csv')]

    if not files:
        print("No CSV files found in the folder.")
        return

    # Write the sentence counts to the output file
    with open(output_file, 'w') as f:
        f.write("File Name, Number\n")

        # Iterate over each file and count the number of sentences
        for file in files:
            file_path = os.path.join(output_folder, file)
            df = pd.read_csv(file_path)

            # Assuming each row represents a sentence
            num_sentences = len(df)

            # Write the result to the file
            f.write(f"{file}, {num_sentences}\n")
            print(f"{file}: {num_sentences} sentences")

def main():
    """
    Main function to orchestrate the workflow of reading, splitting, and processing data.
    """
    input_file = "output_data/personal care.csv"  # Replace with the path to your CSV file
    chunk_size = 2291
    output_files = ["Rohit_benergy_(personal_care_data)"]

    # Read the CSV file
    df = pd.read_csv(input_file)

    # Split and process data into chunks for each intern
    split_and_process_chunks(df, chunk_size, output_files)

    # Count sentences in output files and write to an output CSV
    output_folder = "output_data"  # Replace with your output folder path
    output_file = "sentence_counts.csv"  # The file where results will be stored
    count_sentences_in_files(output_folder, output_file)

if __name__ == "__main__":
    main()
